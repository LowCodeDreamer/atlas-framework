#!/usr/bin/env python3
"""
extract.py - Content extraction per file type

Extracts content from files based on their handler type.
Part of the Atlas file normalization pipeline.
"""

import json
import sys
from pathlib import Path
from typing import Optional
import re

# Optional imports for document handling
try:
    import yaml
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False

try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False


def extract_text(file_path: str) -> dict:
    """
    Extract content from plain text files (.txt, .md, .csv, .log).

    Returns:
        dict with:
            - content: The raw text content
            - metadata: Any extracted metadata (e.g., YAML frontmatter)
            - success: Boolean
            - error: Error message if failed
    """
    try:
        path = Path(file_path)
        content = path.read_text(encoding='utf-8')

        # Check for YAML frontmatter (already normalized files)
        metadata = {}
        if content.startswith('---'):
            parts = content.split('---', 2)
            if len(parts) >= 3:
                try:
                    if YAML_AVAILABLE:
                        metadata = yaml.safe_load(parts[1]) or {}
                    content = parts[2].strip()
                except Exception:
                    pass  # Not valid YAML frontmatter, use full content

        return {
            'content': content,
            'metadata': metadata,
            'char_count': len(content),
            'line_count': content.count('\n') + 1,
            'success': True,
            'error': None
        }
    except Exception as e:
        return {
            'content': None,
            'metadata': {},
            'success': False,
            'error': str(e)
        }


def extract_structured(file_path: str) -> dict:
    """
    Extract content from structured data files (.json, .yaml, .xml).

    Returns dict with parsed content and formatted string representation.
    """
    try:
        path = Path(file_path)
        ext = path.suffix.lower()
        raw_content = path.read_text(encoding='utf-8')

        if ext == '.json':
            data = json.loads(raw_content)
            formatted = json.dumps(data, indent=2)
            return {
                'content': formatted,
                'data': data,
                'format': 'json',
                'success': True,
                'error': None
            }

        elif ext in ['.yaml', '.yml']:
            if not YAML_AVAILABLE:
                return {
                    'content': raw_content,
                    'data': None,
                    'format': 'yaml',
                    'success': True,
                    'error': 'YAML parsing unavailable (pyyaml not installed), using raw content'
                }
            data = yaml.safe_load(raw_content)
            return {
                'content': raw_content,
                'data': data,
                'format': 'yaml',
                'success': True,
                'error': None
            }

        elif ext == '.xml':
            # For XML, just return raw content (proper parsing needs more setup)
            return {
                'content': raw_content,
                'data': None,
                'format': 'xml',
                'success': True,
                'error': None
            }

        else:
            return {
                'content': raw_content,
                'data': None,
                'format': 'unknown',
                'success': True,
                'error': None
            }

    except json.JSONDecodeError as e:
        return {
            'content': None,
            'success': False,
            'error': f'Invalid JSON: {e}'
        }
    except Exception as e:
        return {
            'content': None,
            'success': False,
            'error': str(e)
        }


def extract_pdf(file_path: str) -> dict:
    """
    Extract text content from PDF files.

    Requires PyPDF2 to be installed.
    """
    if not PDF_AVAILABLE:
        return {
            'content': None,
            'success': False,
            'error': 'PDF extraction requires PyPDF2. Install with: pip install PyPDF2',
            'needs_library': 'PyPDF2'
        }

    try:
        text_parts = []
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            page_count = len(reader.pages)

            for page in reader.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)

        content = '\n\n'.join(text_parts)

        return {
            'content': content,
            'page_count': page_count,
            'char_count': len(content),
            'success': True,
            'error': None
        }
    except Exception as e:
        return {
            'content': None,
            'success': False,
            'error': f'PDF extraction failed: {e}'
        }


def extract_docx(file_path: str) -> dict:
    """
    Extract text content from Word documents (.docx).

    Requires python-docx to be installed.
    """
    if not DOCX_AVAILABLE:
        return {
            'content': None,
            'success': False,
            'error': 'DOCX extraction requires python-docx. Install with: pip install python-docx',
            'needs_library': 'python-docx'
        }

    try:
        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        content = '\n\n'.join(paragraphs)

        return {
            'content': content,
            'paragraph_count': len(paragraphs),
            'char_count': len(content),
            'success': True,
            'error': None
        }
    except Exception as e:
        return {
            'content': None,
            'success': False,
            'error': f'DOCX extraction failed: {e}'
        }


def extract_document(file_path: str) -> dict:
    """
    Extract content from document files (.pdf, .docx, .doc, .rtf).

    Routes to appropriate extractor based on extension.
    """
    ext = Path(file_path).suffix.lower()

    if ext == '.pdf':
        return extract_pdf(file_path)
    elif ext == '.docx':
        return extract_docx(file_path)
    elif ext == '.doc':
        return {
            'content': None,
            'success': False,
            'error': 'Legacy .doc format not supported. Convert to .docx first.'
        }
    elif ext == '.rtf':
        # RTF can be read as text with some cleanup
        try:
            content = Path(file_path).read_text(encoding='utf-8', errors='ignore')
            # Strip RTF control codes (basic cleanup)
            content = re.sub(r'\\[a-z]+\d*\s?', '', content)
            content = re.sub(r'[{}]', '', content)
            return {
                'content': content.strip(),
                'success': True,
                'error': 'RTF parsing is basic; some formatting may be lost'
            }
        except Exception as e:
            return {
                'content': None,
                'success': False,
                'error': f'RTF extraction failed: {e}'
            }
    else:
        return {
            'content': None,
            'success': False,
            'error': f'Unsupported document format: {ext}'
        }


def extract_vision(file_path: str) -> dict:
    """
    Prepare image for Vision API analysis.

    Images require Claude's vision capability for content extraction.
    This function returns metadata and flags that AI processing is needed.
    """
    try:
        path = Path(file_path)
        file_size = path.stat().st_size

        return {
            'content': None,
            'file_path': str(path.absolute()),
            'file_size': file_size,
            'requires_vision_api': True,
            'prompt_suggestion': 'Analyze this image and describe its contents. If it contains text, extract it. If it contains diagrams, explain them.',
            'success': True,
            'error': None
        }
    except Exception as e:
        return {
            'content': None,
            'success': False,
            'error': f'Image preparation failed: {e}'
        }


def extract_audio(file_path: str) -> dict:
    """
    Prepare audio file for transcription.

    Audio files require external transcription service.
    """
    try:
        path = Path(file_path)
        file_size = path.stat().st_size

        return {
            'content': None,
            'file_path': str(path.absolute()),
            'file_size': file_size,
            'requires_transcription': True,
            'prompt_suggestion': 'This is an audio file that needs transcription.',
            'success': True,
            'error': None
        }
    except Exception as e:
        return {
            'content': None,
            'success': False,
            'error': f'Audio preparation failed: {e}'
        }


def extract_video(file_path: str) -> dict:
    """
    Prepare video file for processing.

    Video files may need frame extraction and/or audio transcription.
    """
    try:
        path = Path(file_path)
        file_size = path.stat().st_size

        return {
            'content': None,
            'file_path': str(path.absolute()),
            'file_size': file_size,
            'requires_video_processing': True,
            'prompt_suggestion': 'This is a video file. Consider extracting key frames and/or transcribing audio.',
            'success': True,
            'error': None
        }
    except Exception as e:
        return {
            'content': None,
            'success': False,
            'error': f'Video preparation failed: {e}'
        }


def extract_spreadsheet(file_path: str) -> dict:
    """
    Extract content from spreadsheet files (.xlsx, .xls, .ods).

    Basic implementation - returns that openpyxl would be needed.
    """
    try:
        # Try openpyxl if available
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            sheets_data = {}

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    # Filter out completely empty rows
                    if any(cell is not None for cell in row):
                        rows.append([str(cell) if cell is not None else '' for cell in row])
                sheets_data[sheet_name] = rows

            # Format as markdown tables
            content_parts = []
            for sheet_name, rows in sheets_data.items():
                if rows:
                    content_parts.append(f"## {sheet_name}\n")
                    # Header row
                    if rows:
                        content_parts.append('| ' + ' | '.join(rows[0]) + ' |')
                        content_parts.append('| ' + ' | '.join(['---'] * len(rows[0])) + ' |')
                        for row in rows[1:]:
                            content_parts.append('| ' + ' | '.join(row) + ' |')
                    content_parts.append('')

            return {
                'content': '\n'.join(content_parts),
                'sheet_count': len(sheets_data),
                'success': True,
                'error': None
            }

        except ImportError:
            return {
                'content': None,
                'success': False,
                'error': 'Spreadsheet extraction requires openpyxl. Install with: pip install openpyxl',
                'needs_library': 'openpyxl'
            }

    except Exception as e:
        return {
            'content': None,
            'success': False,
            'error': f'Spreadsheet extraction failed: {e}'
        }


# Extractor routing
EXTRACTORS = {
    'text': extract_text,
    'structured': extract_structured,
    'document': extract_document,
    'vision': extract_vision,
    'audio': extract_audio,
    'video': extract_video,
    'spreadsheet': extract_spreadsheet,
    'presentation': lambda f: {'content': None, 'success': False, 'error': 'Presentation extraction not yet implemented'},
}


def extract_content(file_path: str, extractor_type: str) -> dict:
    """
    Extract content using the specified extractor.

    Args:
        file_path: Path to file
        extractor_type: Type of extractor (text, document, vision, etc.)

    Returns:
        Extraction result dict
    """
    if extractor_type not in EXTRACTORS:
        return {
            'content': None,
            'success': False,
            'error': f'Unknown extractor type: {extractor_type}'
        }

    extractor = EXTRACTORS[extractor_type]
    result = extractor(file_path)
    result['extractor'] = extractor_type
    result['file_path'] = file_path

    return result


def main():
    """CLI interface for content extraction."""
    if len(sys.argv) < 3:
        print("Usage: extract.py <file_path> <extractor_type> [--json]", file=sys.stderr)
        print("Extractor types: text, structured, document, vision, audio, video, spreadsheet", file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1]
    extractor_type = sys.argv[2]
    json_output = '--json' in sys.argv

    result = extract_content(file_path, extractor_type)

    if json_output:
        # Don't include full content in JSON output if very long
        output = result.copy()
        if output.get('content') and len(output['content']) > 1000:
            output['content_preview'] = output['content'][:500] + '...'
            output['content_length'] = len(output['content'])
            del output['content']
        print(json.dumps(output, indent=2, default=str))
    else:
        if result['success']:
            print(f"Extractor: {extractor_type}")
            print(f"Success: {result['success']}")
            if result.get('content'):
                preview = result['content'][:500]
                if len(result['content']) > 500:
                    preview += '...'
                print(f"\nContent preview:\n{preview}")
            if result.get('requires_vision_api'):
                print("\n⚠️  This file requires Vision API for content extraction")
            if result.get('requires_transcription'):
                print("\n⚠️  This file requires transcription service")
        else:
            print(f"Error: {result['error']}", file=sys.stderr)
            sys.exit(1)


if __name__ == '__main__':
    main()
